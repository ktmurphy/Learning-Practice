# Write your code here :-)
import openpyxl
from openpyxl.styles import Font

wb = openpyxl.Workbook()
sheet = wb["Sheet"]
italic24font = Font(size=24, italic=True)
sheet["A1"].font = italic24font
sheet["A1"] = "Hello World"
sheet["A2"] = 200
sheet["A3"] = 300
sheet["A4"] = "=SUM(A2:A3)"
sheet["A1"] = "Tall Row"
sheet["B2"] = "Wide Column"
sheet.row_dimensions[1].height = 70
sheet.column_dimensions["B"].width = 20

sheet.merge_cells("C4:F8")
sheet["C4"] = "Merged cells"


sheet.unmerge_cells("C4:F8")

sheet.freeze_panes = "A2"

wb.save("styles.xlsx")
