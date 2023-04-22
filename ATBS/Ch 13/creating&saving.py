# Write your code here :-)
import openpyxl

wb = openpyxl.Workbook()
print("First sheet name list:", wb.sheetnames)
wb.create_sheet()
print("After adding new sheet:", wb.sheetnames)
wb.create_sheet(index=0, title="First Sheet")
print("After adding sheet called 'First Sheet':", wb.sheetnames)
wb.create_sheet(index=2, title="Middle Sheet")
print("After adding sheet called 'Middle Sheet':", wb.sheetnames)
del wb["Middle Sheet"]
del wb["Sheet1"]
print("After deleting some sheets:", wb.sheetnames)
