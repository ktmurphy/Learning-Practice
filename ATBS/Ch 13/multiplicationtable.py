# Write your code here :-)


import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
sheet = wb.active
for i in range(1, 11):
    sheet["A" + str(i + 1)] = i
for i in range(1, 11):
    sheet[get_column_letter(i + 1) + "1"] = i


letter = get_column_letter(sheet.max_column)
row = str(sheet.max_row)

max_cell = str(letter + row)
print(max_cell)
chunk = tuple(sheet["B2" : str(max_cell)])
print(chunk)
for row in chunk:
    for cell in row:
        # product of value in [Column in question in row 1 and  column A row in question
        cell.value = (
            "=PRODUCT(A"
            + str(cell.row)
            + ","
            + str(get_column_letter(cell.column))
            + "1)"
        )


wb.save("multiplicationTable.xlsx")
